import random
from openpyxl import load_workbook

excel_file = "Student_List_2nd_Sem.xlsx"

# Load workbook directly
wb = load_workbook(excel_file)

print(f"Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Find header row (look for "Name" column)
    header_row = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=False), 1):
        row_values = [cell.value for cell in row if cell.value]
        if any("name" in str(v).lower() or "enrollment" in str(v).lower() for v in row_values):
            header_row = row_idx
            break
    
    if header_row == 0:
        continue
    
    print(f"\n✓ {sheet_name}: header at row {header_row}")
    
    # Add email and phone columns after last column
    max_col = ws.max_column + 1
    
    # Add headers
    ws.cell(row=header_row, column=max_col, value="email")
    ws.cell(row=header_row, column=max_col+1, value="phone")
    
    # Add data starting from header_row+1
    data_row = header_row + 1
    row_count = 0
    
    while ws.cell(row=data_row, column=1).value:
        domains = ["gmail.com", "yahoo.com", "outlook.com", "student.edu"]
        name_part = f"student{data_row}{random.randint(10,99)}"
        email = f"{name_part}@{random.choice(domains)}"
        phone = f"9{random.randint(100000000, 999999999)}"
        
        ws.cell(row=data_row, column=max_col, value=email)
        ws.cell(row=data_row, column=max_col+1, value=phone)
        
        data_row += 1
        row_count += 1
    
    print(f"   Added {row_count} emails and phones")

wb.save(excel_file)
print(f"\n✓ Excel updated successfully with contact data!")
